"""Tabular ingestion for CSV and Excel sources."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook

from ..models.domain import Pack, PackItem
from ..utils import ensure_dir, slugify

logger = logging.getLogger(__name__)


COLUMN_ALIASES = {
    "pack": "pack_name",
    "bundle": "pack_name",
    "bundle_name": "pack_name",
    "pack_name": "pack_name",
    "name": "pack_name",
    "event_shop": "event_shop",
    "event shop": "event_shop",
    "shop_type": "shop_type",
    "shop type": "shop_type",
    "price": "price",
    "price_usd": "price",
    "price($)": "price",
    "cost": "price",
    "usd": "price",
    "currency": "currency",
    "$": "currency",
    "item": "item_name",
    "items": "item_name",
    "reward": "item_name",
    "item_name": "item_name",
    "quantity": "quantity",
    "qty": "quantity",
    "amount": "quantity",
    "count": "quantity",
    "category": "category",
    "type": "category",
    "tag": "tags",
    "tags": "tags",
    "note": "notes",
    "gem per unit": "gem_per_unit",
    "gems per unit": "gem_per_unit",
    "gem_value": "gem_value",
    "weighted gem value": "weighted_gem_value",
    "token cost": "token_cost",
    "equivalent gem cost": "equivalent_gem_cost",
}

HEADER_KEYWORDS = {
    "item",
    "items",
    "quantity",
    "qty",
    "gem per unit",
    "gems per unit",
    "gem value",
    "total",
    "price",
    "cost",
    "currency",
    "token",
}


def _normalize_columns(columns: Iterable[str]) -> List[str]:
    normalized = []
    for col in columns:
        key = str(col).strip().lower()
        key = key.replace(" ", "_").replace("(", "").replace(")", "")
        key = COLUMN_ALIASES.get(key, key)
        normalized.append(key)
    return normalized


def _to_float(value) -> float:
    try:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        cleaned = str(value).replace("$", "").replace(",", "").strip()
        return float(cleaned) if cleaned else 0.0
    except Exception:
        return 0.0


def _extract_images(ws, dest_dir: Path, prefix: str) -> List[Tuple[int, Path]]:
    """Extract embedded images and return mapping of row -> saved path."""
    ensure_dir(dest_dir)
    mappings: List[Tuple[int, Path]] = []
    for idx, image in enumerate(getattr(ws, "_images", []), start=1):
        row_num = None
        try:
            row_num = image.anchor._from.row + 1  # type: ignore[attr-defined]
        except Exception:
            row_num = None
        filename = dest_dir / f"{prefix}_img_{idx}.png"
        try:
            data = image._data()  # type: ignore[attr-defined]
            with filename.open("wb") as f:
                f.write(data)
            mappings.append((row_num, filename))
        except Exception as exc:
            logger.warning("Could not extract image %s: %s", filename.name, exc)
    return mappings


def _expand_merged_cells(ws) -> Dict[Tuple[int, int], object]:
    """Return mapping of merged-cell coordinates to their top-left value."""
    merged_values: Dict[Tuple[int, int], object] = {}
    for merged in ws.merged_cells.ranges:
        top_left = merged.start_cell
        value = top_left.value
        for row in ws[merged.coord]:
            for cell in row:
                merged_values[(cell.row, cell.column)] = value
    return merged_values


def _header_score(row: Sequence[Optional[object]]) -> int:
    tokens = []
    for val in row:
        if val is None:
            continue
        text = str(val).strip().lower()
        if not text:
            continue
        tokens.append(text)
    score = 0
    for t in tokens:
        for kw in HEADER_KEYWORDS:
            if kw in t:
                score += 1
    return score


def _is_reference(sheet_name: str, pack_hint: Optional[str], config: Dict | None) -> bool:
    patterns = []
    if config:
        patterns = [p.lower() for p in config.get("sheet_name_patterns", []) if isinstance(p, str)]
    for value in [sheet_name, pack_hint]:
        if not value:
            continue
        low = value.lower()
        if any(pat in low for pat in patterns):
            return True
    return False


def _sheet_tables(ws, sheet_name: str, reference_config: Dict | None) -> List[Tuple[pd.DataFrame, List[int], Optional[str], bool]]:
    """Split a worksheet into multiple logical tables based on header detection."""
    merged_values = _expand_merged_cells(ws)
    rows = []
    for excel_row in ws.iter_rows():
        values = []
        for cell in excel_row:
            val = cell.value
            if val is None and (cell.row, cell.column) in merged_values:
                val = merged_values[(cell.row, cell.column)]
            values.append(val)
        rows.append(values)
    tables: List[Tuple[pd.DataFrame, List[int], Optional[str], bool]] = []
    pack_hint: Optional[str] = None
    i = 0
    while i < len(rows):
        row = rows[i]
        values = [v for v in row if v not in (None, "")]
        if not values:
            i += 1
            continue
        # capture pack title rows (single text cell)
        if len(values) == 1 and isinstance(values[0], str):
            pack_hint = str(values[0]).strip()
            i += 1
            continue

        if _header_score(row) >= 2:
            header = row
            data: List[List] = []
            row_numbers: List[int] = []
            empty_run = 0
            i += 1
            while i < len(rows):
                next_row = rows[i]
                if not any(next_row):
                    empty_run += 1
                    if empty_run >= 2:
                        i += 1
                        break
                    i += 1
                    continue
                if _header_score(next_row) >= 2:
                    break
                empty_run = 0
                data.append(list(next_row))
                row_numbers.append(i + 1)
                i += 1

            if data:
                header_vals = [str(h) if h is not None else f"col_{idx+1}" for idx, h in enumerate(header)]
                df = pd.DataFrame(data, columns=_normalize_columns(header_vals))
                is_ref = _is_reference(sheet_name, pack_hint, reference_config)
                tables.append((df, row_numbers, pack_hint, is_ref))
            pack_hint = None
            continue
        i += 1
    return tables


def _normalize_dataframe(df: pd.DataFrame, default_pack_name: str, default_currency: str) -> pd.DataFrame:
    df = df.copy()
    df.columns = _normalize_columns(df.columns)
    if "pack_name" not in df.columns:
        df["pack_name"] = None
    if "event_shop" in df.columns:
        df["pack_name"] = df["pack_name"].fillna(df["event_shop"])
        df["pack_name"] = df["pack_name"].where(df["pack_name"] != "", df["event_shop"])
    if "event_shop" not in df.columns and "shop_type" in df.columns:
        df["pack_name"] = df["pack_name"].fillna(df["shop_type"])
        df["pack_name"] = df["pack_name"].where(df["pack_name"] != "", df["shop_type"])
    df["pack_name"] = df["pack_name"].fillna(default_pack_name).replace("", default_pack_name)
    if "item_name" not in df.columns:
        df["item_name"] = "Unknown Item"
    if "quantity" not in df.columns:
        df["quantity"] = 0
    if "price" not in df.columns:
        df["price"] = 0.0
    if "currency" not in df.columns:
        df["currency"] = default_currency
    if "tags" not in df.columns:
        df["tags"] = ""
    return df


SUMMARY_KEYS = {
    "gem total": "gem_total",
    "pack %": "pack_pct",
    "true pack value %": "true_pack_value_pct",
}

IGNORED_ITEMS = {"gem total", "pack %", "true pack value %", "exclude resources"}


def _pack_from_rows(
    df: pd.DataFrame,
    row_numbers: List[int],
    source_file: Path,
    sheet_name: str | None,
    image_map: Dict[int, Path],
    pack_name_hint: Optional[str] = None,
    is_reference: bool = False,
) -> List[Pack]:
    packs: Dict[str, Pack] = {}
    for idx, (_, row) in enumerate(df.iterrows()):
        pack_name = str(row.get("pack_name") or pack_name_hint or "Unnamed Pack").strip()
        price = _to_float(row.get("price"))
        currency = str(row.get("currency") or "USD").strip().upper()
        tags_val = row.get("tags")
        tags = [t.strip() for t in str(tags_val).split(",") if str(tags_val) and t.strip()] if tags_val else []
        pack_id = slugify(f"{pack_name}-{price}-{source_file.stem}")
        pack = packs.get(pack_id)
        if pack is None:
            pack = Pack(
                pack_id=pack_id,
                name=pack_name,
                price=price,
                currency=currency,
                source_file=str(source_file),
                source_sheet=sheet_name,
                is_reference=is_reference,
                tags=tags,
                items=[],
            )
            packs[pack_id] = pack

        item_name = str(row.get("item_name") or "Unknown Item").strip()
        lowered_item = item_name.lower()
        if lowered_item in SUMMARY_KEYS:
            field = SUMMARY_KEYS[lowered_item]
            pack.meta[field] = _to_float(row.get("total") or row.get("equivalent_gem_cost"))
            continue
        if lowered_item in IGNORED_ITEMS:
            continue

        # skip rows without meaningful quantity or name
        category = str(row.get("category") or "unknown").strip().lower() or "unknown"
        quantity = _to_float(row.get("quantity"))
        item_id = slugify(item_name)
        sheet_row_number = row_numbers[idx] if idx < len(row_numbers) else None
        icon_path = image_map.get(sheet_row_number) if sheet_row_number is not None else None

        base_value = None
        if "gem_per_unit" in row:
            base_value = _to_float(row.get("gem_per_unit"))
        elif "gem_value" in row:
            base_value = _to_float(row.get("gem_value"))
        elif "weighted_gem_value" in row:
            base_value = _to_float(row.get("weighted_gem_value"))
        elif "equivalent_gem_cost" in row and quantity:
            base_value = _to_float(row.get("equivalent_gem_cost")) / quantity

        meta: Dict[str, object] = {}
        if "token_cost" in row:
            meta["token_cost"] = _to_float(row.get("token_cost"))
        if "equivalent_gem_cost" in row:
            meta["equivalent_gem_cost"] = _to_float(row.get("equivalent_gem_cost"))
        if "total" in row and row.get("total") not in (None, ""):
            meta["row_total"] = _to_float(row.get("total"))

        pack.items.append(
            PackItem(
                item_id=item_id,
                name=item_name,
                category=category,
                quantity=quantity,
                icon=str(icon_path) if icon_path else None,
                source_row=sheet_row_number,
                base_value=base_value if base_value not in (None, 0) else None,
                meta=meta,
            )
        )
    return list(packs.values())


def parse_csv(path: Path, default_currency: str = "USD") -> List[Pack]:
    logger.info("Ingesting CSV %s", path.name)
    df = pd.read_csv(path)
    df = _normalize_dataframe(df, default_pack_name=path.stem, default_currency=default_currency)
    row_numbers = list(range(2, len(df) + 2))  # pretend header on row 1 for consistency
    return _pack_from_rows(df, row_numbers, path, None, image_map={})


def parse_excel(
    path: Path,
    images_dir: Path,
    default_currency: str = "USD",
    reference_config: Dict | None = None,
) -> List[Pack]:
    logger.info("Ingesting Excel %s", path.name)
    workbook = load_workbook(path, data_only=True)
    all_packs: List[Pack] = []
    for sheet_name in workbook.sheetnames:
        if "instruction" in sheet_name.lower():
            continue
        ws = workbook[sheet_name]
        image_pairs = _extract_images(ws, images_dir, f"{path.stem}_{slugify(sheet_name)}")
        image_map = {row: file for row, file in image_pairs if row is not None}
        tables = _sheet_tables(ws, sheet_name, reference_config)
        for idx, (df, row_numbers, pack_hint, is_ref) in enumerate(tables, start=1):
            if df.empty:
                continue
            default_pack_name = pack_hint or f"{path.stem}-{slugify(sheet_name)}-table-{idx}"
            df = _normalize_dataframe(df, default_pack_name=default_pack_name, default_currency=default_currency)
            sheet_packs = _pack_from_rows(
                df, row_numbers, path, sheet_name, image_map, pack_name_hint=pack_hint, is_reference=is_ref
            )
            all_packs.extend(sheet_packs)
    return all_packs


def parse_file(
    path: Path,
    images_dir: Path,
    default_currency: str = "USD",
    reference_config: Dict | None = None,
) -> List[Pack]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        return parse_csv(path, default_currency=default_currency)
    if suffix in {".xlsx", ".xlsm"}:
        return parse_excel(path, images_dir=images_dir, default_currency=default_currency, reference_config=reference_config)
    logger.warning("Skipping unsupported file: %s", path.name)
    return []
